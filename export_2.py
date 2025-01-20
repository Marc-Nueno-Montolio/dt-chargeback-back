from openpyxl import Workbook
import csv
import json
from settings import LOG_FORMAT, LOG_LEVEL
from typing import Dict
import logging
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
logger = logging.getLogger(__name__)

from datetime import datetime


class ChargebackExcelExporter:
    
    def __init__(self):
        pass

    def export_to_excel(self, report: Dict, output: str):
        """
        Export chargeback report to Excel with flattened entity view
        
        Args:
            report (Dict): Dictionary containing the report data with DGs, IS and entities
            output (str): Output file path for the Excel file
        """
        try:
            # Create empty lists to store data
            data = []
            
            # Define headers
            headers = [
                'DG', 'IS', 'IS Managed', 'Entity Type', 'Entity Name', 'DT ID', 'Managed', 'Cloud', 'Billable', 'Tagged DGs',
                'Fullstack', 'Infra', 'RUM', 'RUM with SR', 'Browser Monitor',
                'HTTP Monitor', '3rd Party Monitor', 'Managed Hosts in IS', 'Non-Managed Hosts in IS'
            ]

            # Sort DGs to put DIGIT C at the end
            sorted_dgs = sorted(report['dgs'], key=lambda x: x['name'] == 'DIGIT C')

            # Create a dictionary to store DG totals
            dg_totals = {}

            # Initialize DIGIT C totals even if not in report
            dg_totals['DIGIT C'] = {
                'fullstack': 0,
                'infra': 0,
                'rum': 0,
                'rum_sr': 0,
                'browser': 0,
                'http': 0,
                'third': 0,
                'managed': 0
            }

            # Iterate through DGs and collect data
            for dg in sorted_dgs:
                dg_name = dg['name']
                logger.info(f"Processing DG: {dg_name}")
                if dg_name not in dg_totals:
                    dg_totals[dg_name] = {
                        'fullstack': 0,
                        'infra': 0,
                        'rum': 0,
                        'rum_sr': 0,
                        'browser': 0,
                        'http': 0,
                        'third': 0,
                        'managed': 0
                    }
                
                # Process assigned entities
                for is_system in dg['data']['information_systems']:
                    is_name = is_system['name']
                    is_managed = is_system.get('managed', False)
                    
                    # Calculate managed and non-managed hosts counts for IS
                    managed_hosts_count = sum(1 for h in is_system['data']['entities'].get('hosts', []) if h.get('managed', False))
                    non_managed_hosts_count = sum(1 for h in is_system['data']['entities'].get('hosts', []) if not h.get('managed', False))
                    
                    # Process hosts
                    for host in is_system['data']['entities'].get('hosts', []):
                        # Sort tagged DGs to put DIGIT C last
                        tagged_dgs = sorted(host.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                        
                        # Determine which DG to attribute costs to
                        target_dg = 'DIGIT C' if not host.get('managed', False) and not host.get('billed', False) else dg_name
                        
                        fullstack_usage = host.get('usage', {}).get('fullstack', 0)
                        infra_usage = host.get('usage', {}).get('infra', 0)
                        
                        # Add usage to appropriate DG totals
                        dg_totals[target_dg]['fullstack'] += fullstack_usage
                        dg_totals[target_dg]['infra'] += infra_usage
                        if host.get('managed', False):
                            dg_totals[target_dg]['managed'] += 1
                            
                        row = [
                            target_dg, is_name, is_managed, 'Host', host.get('name', ''), host.get('dt_id', ''),
                            host.get('managed', False), host.get('cloud', False), host.get('billed', False),
                            ', '.join(tagged_dgs), fullstack_usage, infra_usage, 0, 0, 0, 0, 0,
                            managed_hosts_count, non_managed_hosts_count
                        ]
                        data.append(row)

                    # Process applications
                    for app in is_system['data']['entities'].get('applications', []):
                        target_dg = 'DIGIT C' if not app.get('billed', False) else dg_name
                        tagged_dgs = sorted(app.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                        
                        rum_usage = app.get('usage', {}).get('rum', 0)
                        rum_sr_usage = app.get('usage', {}).get('rum_with_sr', 0)
                        
                        dg_totals[target_dg]['rum'] += rum_usage
                        dg_totals[target_dg]['rum_sr'] += rum_sr_usage
                        
                        row = [
                            target_dg, is_name, is_managed, 'Application', app.get('name', ''), app.get('dt_id', ''),
                            False, None, app.get('billed', False), ', '.join(tagged_dgs), 0, 0,
                            rum_usage, rum_sr_usage, 0, 0, 0, managed_hosts_count, non_managed_hosts_count
                        ]
                        data.append(row)

                    # Process synthetics
                    for synthetic in is_system['data']['entities'].get('synthetics', []):
                        target_dg = 'DIGIT C' if not synthetic.get('billed', False) else dg_name
                        tagged_dgs = sorted(synthetic.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                        
                        browser_usage = synthetic.get('usage', {}).get('browser_monitor', 0)
                        http_usage = synthetic.get('usage', {}).get('http_monitor', 0)
                        third_usage = synthetic.get('usage', {}).get('third_party_monitor', 0)
                        
                        dg_totals[target_dg]['browser'] += browser_usage
                        dg_totals[target_dg]['http'] += http_usage
                        dg_totals[target_dg]['third'] += third_usage
                        
                        row = [
                            target_dg, is_name, is_managed, 'Synthetic', synthetic.get('name', ''), synthetic.get('dt_id', ''),
                            False, None, synthetic.get('billed', False), ', '.join(tagged_dgs), 0, 0,
                            0, 0, browser_usage, http_usage, third_usage,
                            managed_hosts_count, non_managed_hosts_count
                        ]
                        data.append(row)

                # Process unassigned entities similarly
                unassigned = dg['data']['unassigned_entities']['entities']
                
                # Calculate managed and non-managed hosts counts for unassigned
                unassigned_managed_hosts_count = sum(1 for h in unassigned.get('hosts', []) if h.get('managed', False))
                unassigned_non_managed_hosts_count = sum(1 for h in unassigned.get('hosts', []) if not h.get('managed', False))

                # Add unassigned entities
                for entity_type, entities in unassigned.items():
                    if entity_type == 'hosts':
                        for host in entities:
                            target_dg = 'DIGIT C' if not host.get('managed', False) and not host.get('billed', False) else dg_name
                            tagged_dgs = sorted(host.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                            
                            fullstack_usage = host.get('usage', {}).get('fullstack', 0)
                            infra_usage = host.get('usage', {}).get('infra', 0)
                            
                            dg_totals[target_dg]['fullstack'] += fullstack_usage
                            dg_totals[target_dg]['infra'] += infra_usage
                            if host.get('managed', False):
                                dg_totals[target_dg]['managed'] += 1
                                
                            row = [
                                target_dg, 'Unassigned', False, 'Host', host.get('name', ''), host.get('dt_id', ''),
                                host.get('managed', False), host.get('cloud', False), host.get('billed', False),
                                ', '.join(tagged_dgs), fullstack_usage, infra_usage, 0, 0, 0, 0, 0,
                                unassigned_managed_hosts_count, unassigned_non_managed_hosts_count
                            ]
                            data.append(row)
                    elif entity_type == 'applications':
                        for app in entities:
                            target_dg = 'DIGIT C' if not app.get('billed', False) else dg_name
                            tagged_dgs = sorted(app.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                            
                            rum_usage = app.get('usage', {}).get('rum', 0)
                            rum_sr_usage = app.get('usage', {}).get('rum_with_sr', 0)
                            
                            dg_totals[target_dg]['rum'] += rum_usage
                            dg_totals[target_dg]['rum_sr'] += rum_sr_usage
                            
                            row = [
                                target_dg, 'Unassigned', False, 'Application', app.get('name', ''), app.get('dt_id', ''),
                                False, None, app.get('billed', False), ', '.join(tagged_dgs), 0, 0,
                                rum_usage, rum_sr_usage, 0, 0, 0,
                                unassigned_managed_hosts_count, unassigned_non_managed_hosts_count
                            ]
                            data.append(row)
                    elif entity_type == 'synthetics':
                        for synthetic in entities:
                            target_dg = 'DIGIT C' if not synthetic.get('billed', False) else dg_name
                            tagged_dgs = sorted(synthetic.get('tagged_dgs', []), key=lambda x: x == 'DIGIT C')
                            
                            browser_usage = synthetic.get('usage', {}).get('browser_monitor', 0)
                            http_usage = synthetic.get('usage', {}).get('http_monitor', 0)
                            third_usage = synthetic.get('usage', {}).get('third_party_monitor', 0)
                            
                            dg_totals[target_dg]['browser'] += browser_usage
                            dg_totals[target_dg]['http'] += http_usage
                            dg_totals[target_dg]['third'] += third_usage
                            
                            row = [
                                target_dg, 'Unassigned', False, 'Synthetic', synthetic.get('name', ''), synthetic.get('dt_id', ''),
                                False, None, synthetic.get('billed', False), ', '.join(tagged_dgs), 0, 0,
                                0, 0, browser_usage, http_usage, third_usage,
                                unassigned_managed_hosts_count, unassigned_non_managed_hosts_count
                            ]
                            data.append(row)

            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Create Excel writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Chargeback Breakdown', index=False)
                
                # Get the workbook and worksheet
                wb = writer.book
                ws = wb['Chargeback Breakdown']

                # Define custom column widths
                column_widths = {
                    'A': 15,  # DG
                    'B': 30,  # IS
                    'C': 12,  # IS Managed
                    'D': 12,  # Entity Type
                    'E': 40,  # Entity Name
                    'F': 15,  # DT ID
                    'G': 10,  # Managed
                    'H': 10,  # Cloud
                    'I': 10,  # Billable
                    'J': 30,  # Tagged DGs
                    'K': 10,  # Fullstack
                    'L': 10,  # Infra
                    'M': 10,  # RUM
                    'N': 12,  # RUM with SR
                    'O': 15,  # Browser Monitor
                    'P': 15,  # HTTP Monitor
                    'Q': 15,  # 3rd Party Monitor
                    'R': 20,  # Managed Hosts in IS
                    'S': 20   # Non-Managed Hosts in IS
                }

                # Apply custom column widths
                for col_letter, width in column_widths.items():
                    ws.column_dimensions[col_letter].width = width

                # Create table
                tab = Table(displayName="ChargebackTable", ref=ws.dimensions)
                
                # Add a custom style with dark header and white text
                style = TableStyleInfo(
                    name="TableStyleMedium2", 
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=True, 
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                
                # Add the table to the worksheet
                ws.add_table(tab)

                # Style the header row with dark background and white text
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Freeze the header row
                ws.freeze_panes = 'A2'
                # Create Summary sheet using report data
                summary_data = []
                total_fullstack = 0
                total_infra = 0
                total_rum = 0
                total_rum_sr = 0
                total_browser = 0
                total_http = 0
                total_third = 0
                total_managed = 0

                # Sort DGs for summary, ensuring DIGIT C is last
                sorted_dgs = sorted(report['dgs'], key=lambda x: x['name'] == 'DIGIT C')

                logger.info("Processing DG totals for summary sheet")
                for dg in sorted_dgs:
                    dg_name = dg['name']
                    logger.info(f"Processing totals for DG: {dg_name}")
                    
                    totals = dg['data']['totals']['usage']
                    managed_hosts = dg['data']['totals']['managed_hosts']

                    logger.debug(f"DG {dg_name} raw totals: {totals}")

                    fullstack = totals['fullstack']
                    infra = totals['infra'] 
                    rum = totals['rum']
                    rum_sr = totals['rum_with_sr']
                    browser = totals['browser_monitor']
                    http = totals['http_monitor']
                    third = totals['3rd_party_monitor']

                    total_fullstack += fullstack
                    total_infra += infra
                    total_rum += rum
                    total_rum_sr += rum_sr
                    total_browser += browser
                    total_http += http
                    total_third += third
                    total_managed += managed_hosts

                    logger.debug(f"DG {dg_name} processed totals: Fullstack={fullstack}, Infra={infra}, RUM={rum}, RUM SR={rum_sr}, Browser={browser}, HTTP={http}, Third={third}, Managed={managed_hosts}")

                    summary_data.append([
                        dg_name,
                        fullstack,
                        infra,
                        rum,
                        rum_sr,
                        browser,
                        http,
                        third,
                        managed_hosts
                    ])

                # Add DIGIT C totals if not already included
                if 'DIGIT C' not in [row[0] for row in summary_data]:
                    digit_c_totals = dg_totals['DIGIT C']
                    summary_data.append([
                        'DIGIT C',
                        digit_c_totals['fullstack'],
                        digit_c_totals['infra'],
                        digit_c_totals['rum'],
                        digit_c_totals['rum_sr'],
                        digit_c_totals['browser'],
                        digit_c_totals['http'],
                        digit_c_totals['third'],
                        digit_c_totals['managed']
                    ])

                logger.info(f"Final totals: Fullstack={total_fullstack}, Infra={total_infra}, RUM={total_rum}, RUM SR={total_rum_sr}, Browser={total_browser}, HTTP={total_http}, Third={total_third}, Managed={total_managed}")

                summary_headers = [
                    'DG', 'Fullstack', 'Infra', 'RUM', 'RUM with SR',
                    'Browser Monitor', 'HTTP Monitor', '3rd Party Monitor',
                    'Total Managed Hosts'
                ]

                summary_df = pd.DataFrame(summary_data, columns=summary_headers)
                summary_df.to_excel(writer, sheet_name='DG Summary', index=False)

                # Get the summary worksheet
                summary_ws = wb['DG Summary']

                # Create table for summary with matching style
                summary_tab = Table(displayName="SummaryTable", ref=summary_ws.dimensions)
                summary_tab.tableStyleInfo = style
                summary_ws.add_table(summary_tab)

                # Style the summary header row to match
                for cell in summary_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Add totals row after a blank line
                last_row = len(summary_data) + 2  # +2 for header and blank line
                total_row = last_row + 1


                # Style totals and units
                bold_font = Font(bold=True)
                for col in range(1, 10):
                    cell = summary_ws.cell(row=total_row, column=col)
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center')
                    

                # Add totals row after a blank line
                last_row = len(summary_data) + 2  # +2 for header and blank line
                total_row = last_row + 1

                # Set column widths for summary sheet
                summary_ws.column_dimensions['A'].width = 20  # DG name
                for col in range(ord('B'), ord('I')+1):
                    summary_ws.column_dimensions[chr(col)].width = 15

            logger.info(f"Excel report successfully exported to {output}")
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise