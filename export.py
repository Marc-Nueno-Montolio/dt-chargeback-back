from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import csv
import json
from settings import LOG_FORMAT, LOG_LEVEL
from typing import Dict
import logging
import pandas as pd

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
logger = logging.getLogger(__name__)

from datetime import datetime

def export_data(data, output, format):
    if format == 'json':
        # Convert datetime objects to strings
        for item in data:
            for key, value in item.items():
                if isinstance(value, datetime):
                    item[key] = value.isoformat()
        
        with open(output, 'w') as f:
            json.dump(data, f, indent=2)
    elif format == 'csv':
        with open(output, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(data[0].keys())
            for row in data:
                writer.writerow(row.values())


class ChargebackExcelExporter:
    """
    Handles the export of chargeback report data to Excel format with multiple sheets
    and formatted tables.
    """

    def __init__(self):
        # Define usage columns in order
        self.host_columns = ['Fullstack', 'Infrastructure'] 
        self.app_columns = ['RUM', 'RUM with Session Replay']
        self.synthetic_columns = ['Browser Monitor', 'HTTP Monitor', '3rd Party Monitor']

        # Define styles
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        self.host_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        self.app_fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
        self.synthetic_fill = PatternFill(start_color='F2E7F2', end_color='F2E7F2', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True)
        self.subheader_font = Font(bold=True)
        self.total_font = Font(bold=True)

    def export_to_excel(self, report: Dict, output_path: str):
        """
        Creates an Excel workbook with summary and detailed DG sheets.
        
        Args:
            report: The chargeback report dictionary
            output_path: Path where the Excel file should be saved
        """
        logger.info(f"Starting Excel export to {output_path}")
        
        workbook = Workbook()
        
        # Create summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = 'Summary'
        self._create_summary_sheet(report, summary_sheet)
        
        # Create individual DG sheets
        for dg in report['dgs']:
            sheet_name = dg['name'][:31]  # Excel sheet names limited to 31 chars
            dg_sheet = workbook.create_sheet(sheet_name)
            self._create_dg_sheet(dg, dg_sheet)
            
        # Save workbook
        workbook.save(output_path)
        logger.info("Excel export completed successfully")

    def _create_summary_sheet(self, report: Dict, sheet):
        """Creates the summary sheet with DG totals."""
        logger.debug("Creating summary sheet")
        
        # Add headers
        headers = ['DG'] + self.host_columns + self.app_columns + self.synthetic_columns
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='left')
            
        # Add DG data
        current_row = 2
        for dg in report['dgs']:
            sheet.cell(row=current_row, column=1, value=dg['name'])
            
            # Add usage data
            usage = dg['data']['totals']['usage']
            col = 2
            for usage_type in self.host_columns + self.app_columns + self.synthetic_columns:
                value = usage.get(usage_type.lower().replace(' ', '_'), 0)
                sheet.cell(row=current_row, column=col, value=value)
                col += 1
                
            current_row += 1
            
        # Add totals row
        sheet.cell(row=current_row, column=1, value='TOTAL').font = self.total_font
        usage = report['totals']['usage']
        col = 2
        for usage_type in self.host_columns + self.app_columns + self.synthetic_columns:
            value = usage.get(usage_type.lower().replace(' ', '_'), 0)
            cell = sheet.cell(row=current_row, column=col, value=value)
            cell.font = self.total_font
            col += 1
            
        self._apply_formatting(sheet)

    def _create_dg_sheet(self, dg: Dict, sheet):
        """Creates a detailed sheet for a single DG."""
        logger.debug(f"Creating detail sheet for DG: {dg['name']}")
        
        current_row = 1
        
        # Add DG header and freeze it
        cell = sheet.cell(row=current_row, column=1, value=f"Directorate General: {dg['name']}")
        cell.font = self.header_font
        cell.fill = self.header_fill
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        sheet.freeze_panes = 'A2'  # Freeze the DG header row
        current_row += 1

        # Add main headers and freeze them
        headers = ['', 'Entity Name', 'DT ID', 'Managed', 'Billed', 'Fullstack', 'Infrastructure', 'RUM', 'RUM with Session Replay', 'Browser Monitor', 'HTTP Monitor', '3rd Party Monitor']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        sheet.freeze_panes = f'A{current_row + 1}'  # Freeze both DG header and column headers
        current_row += 2
        
        # Process each IS
        for is_system in dg['data']['information_systems']:
            # Add IS header
            cell = sheet.cell(row=current_row, column=1, value=f"IS: {is_system['name']}  {"(managed)" if {is_system['managed']} == True else ""}")
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
            current_row += 1
            
            # Add entity sections
            # Add hosts section if exists
            if is_system['data']['entities'].get('hosts'):
                cell = sheet.cell(row=current_row, column=2, value="Hosts")
                cell.font = self.subheader_font
                cell.fill = self.host_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for host in is_system['data']['entities']['hosts']:
                    sheet.cell(row=current_row, column=2, value=host.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=host.get('dt_id', ''))
                    sheet.cell(row=current_row, column=4, value=host.get('managed', False))
                    sheet.cell(row=current_row, column=5, value=host.get('billed', False))
                    sheet.cell(row=current_row, column=6, value=host.get('usage', {}).get('fullstack', 0))
                    sheet.cell(row=current_row, column=7, value=host.get('usage', {}).get('infra', 0))
                    for col in range(8, 12):  # Fill remaining columns with '
                        sheet.cell(row=current_row, column=col, value='')
                    current_row += 1
                current_row += 1
                
            # Add applications section if exists
            if is_system['data']['entities'].get('applications'):
                cell = sheet.cell(row=current_row, column=2, value="Applications")
                cell.font = self.subheader_font
                cell.fill = self.app_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for app in is_system['data']['entities']['applications']:
                    sheet.cell(row=current_row, column=2, value=app.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=app.get('dt_id', ''))
                    sheet.cell(row=current_row, column=4, value=False)
                    sheet.cell(row=current_row, column=5, value=app.get('billed', False))
                    sheet.cell(row=current_row, column=6, value='')  # Fullstack
                    sheet.cell(row=current_row, column=7, value='')  # Infrastructure
                    sheet.cell(row=current_row, column=8, value=app.get('usage', {}).get('rum', 0))
                    sheet.cell(row=current_row, column=9, value=app.get('usage', {}).get('rum_with_session_replay', 0))
                    sheet.cell(row=current_row, column=10, value='')
                    sheet.cell(row=current_row, column=11, value='')
                    sheet.cell(row=current_row, column=12, value='')
                    current_row += 1
                current_row += 1
                
            # Add synthetics section if exists
            if is_system['data']['entities'].get('synthetics'):
                cell = sheet.cell(row=current_row, column=2, value="Synthetics")
                cell.font = self.subheader_font
                cell.fill = self.synthetic_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for synthetic in is_system['data']['entities']['synthetics']:
                    sheet.cell(row=current_row, column=2, value=synthetic.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=synthetic.get('dt_id', ''))
                    sheet.cell(row=current_row, column=4, value=False)
                    sheet.cell(row=current_row, column=5, value=synthetic.get('billed', False))
                    sheet.cell(row=current_row, column=6, value='')
                    sheet.cell(row=current_row, column=7, value='')
                    sheet.cell(row=current_row, column=8, value='')
                    sheet.cell(row=current_row, column=9, value='')
                    sheet.cell(row=current_row, column=10, value=synthetic.get('usage', {}).get('browser_monitor', 0))
                    sheet.cell(row=current_row, column=11, value=synthetic.get('usage', {}).get('http_monitor', 0))
                    sheet.cell(row=current_row, column=12, value=synthetic.get('usage', {}).get('third_party_monitor', 0))
                    current_row += 1
                current_row += 1
            
        # Add unassigned entities if present
        if any(len(dg['data']['unassigned_entities']['entities'][entity_type]) > 0 
               for entity_type in ['hosts', 'applications', 'synthetics']):
            
            cell = sheet.cell(row=current_row, column=1, value="Unassigned Entities")
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            current_row += 1
            
            # Add hosts section if exists
            if dg['data']['unassigned_entities']['entities']['hosts']:
                cell = sheet.cell(row=current_row, column=2, value="Hosts")
                cell.font = self.subheader_font
                cell.fill = self.host_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for host in dg['data']['unassigned_entities']['entities']['hosts']:
                    sheet.cell(row=current_row, column=2, value=host.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=host.get('dt_id', ''))
                    sheet.cell(row=current_row, column=4, value=host.get('managed', False))
                    sheet.cell(row=current_row, column=5, value=host.get('billed', False))
                    sheet.cell(row=current_row, column=6, value=host.get('usage', {}).get('fullstack', 0))
                    sheet.cell(row=current_row, column=7, value=host.get('usage', {}).get('infra', 0))
                    for col in range(8, 12):  # Fill remaining columns with empty values
                        sheet.cell(row=current_row, column=col, value='')
                    current_row += 1
                current_row += 1
                
            # Add applications section if exists
            if dg['data']['unassigned_entities']['entities']['applications']:
                cell = sheet.cell(row=current_row, column=2, value="Applications") 
                cell.font = self.subheader_font
                cell.fill = self.app_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for app in dg['data']['unassigned_entities']['entities']['applications']:
                    sheet.cell(row=current_row, column=2, value=app.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=app.get('dt_id', ''))
                    sheet.cell(row=current_row, column=4, value='')
                    sheet.cell(row=current_row, column=5, value='')
                    sheet.cell(row=current_row, column=6, value='')
                    sheet.cell(row=current_row, column=7, value='')
                    sheet.cell(row=current_row, column=8, value=app.get('usage', {}).get('rum', 0))
                    sheet.cell(row=current_row, column=9, value=app.get('usage', {}).get('rum_with_session_replay', 0))
                    sheet.cell(row=current_row, column=10, value='')
                    sheet.cell(row=current_row, column=11, value='')
                    sheet.cell(row=current_row, column=12, value='')
                    current_row += 1
                current_row += 1
                
            # Add synthetics section if exists
            if dg['data']['unassigned_entities']['entities']['synthetics']:
                cell = sheet.cell(row=current_row, column=2, value="Synthetics")
                cell.font = self.subheader_font
                cell.fill = self.synthetic_fill
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
                current_row += 1
                
                for synthetic in dg['data']['unassigned_entities']['entities']['synthetics']:
                    sheet.cell(row=current_row, column=2, value=synthetic.get('name', ''))
                    sheet.cell(row=current_row, column=3, value=synthetic.get('dt_id', ''))
                    for col in range(4, 9):  # Fill host and app columns with empty values
                        sheet.cell(row=current_row, column=col, value='')
                    sheet.cell(row=current_row, column=10, value=synthetic.get('usage', {}).get('browser_monitor', 0))
                    sheet.cell(row=current_row, column=11, value=synthetic.get('usage', {}).get('http_monitor', 0))
                    sheet.cell(row=current_row, column=12, value=synthetic.get('usage', {}).get('third_party_monitor', 0))
                    current_row += 1
                current_row += 1
        
        # Adjust column widths
        sheet.column_dimensions[get_column_letter(col)].width = 2
        for col in range(2, 12):
            sheet.column_dimensions[get_column_letter(col)].width = 15
            
        self._apply_formatting(sheet)

    def _apply_formatting(self, sheet):
        """Applies common formatting to worksheet."""
        # Apply alignment
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                
        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 30)  # Cap width at 30 characters
            sheet.column_dimensions[get_column_letter(cell.column)].width = adjusted_width
